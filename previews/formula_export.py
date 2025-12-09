# Formula Export
import openpyxl
import io
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                             QTableWidget, QTableWidgetItem, QComboBox, QLabel,
                             QFileDialog, QMessageBox, QHeaderView)
from PyQt6.QtCore import Qt, QDate
import pandas as pd
from datetime import datetime

from openpyxl.styles import Side, Border, Alignment, PatternFill, Font
from db import db_call
from utils.send_email import send_email_with_excel_bytes


class ExportPreviewDialog(QDialog):
    def __init__(self, parent, date_from, date_to):
        super().__init__(parent)
        self.parent_widget = parent
        self.date_from = date_from
        self.date_to = date_to
        self.filtered_data = None
        self.headers = ["uid", "Date", "Customer", "Product Code", "Mat Code", "Con", "Deleted"]
        self.excel_bytes = None

        self.setWindowTitle("Export Preview")
        self.setMinimumSize(900, 600)
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout()

        # Filter section
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Filter by Month:"))

        self.month_combo = QComboBox()
        self.populate_months()
        self.month_combo.currentIndexChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.month_combo)

        filter_layout.addStretch()

        email_btn = QPushButton("Send to Email", objectName="PrimaryButton")
        email_btn.clicked.connect(self.send_to_email)
        filter_layout.addWidget(email_btn)

        layout.addLayout(filter_layout)

        # Table preview
        self.table = QTableWidget()
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)

        # Allow manual resizing
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # Set minimum width for all columns
        for col in range(len(self.headers)):
            self.table.horizontalHeader().setMinimumSectionSize(80)

        # Set initial widths
        default_width = 50
        for col in range(len(self.headers)):
            if col == 0:  # Customer column (index 0)
                self.table.setColumnWidth(col, 15)
            elif col == 1:  # Customer column (index 1)
                self.table.setColumnWidth(col, 90)
            elif col == 2:  # Customer column (index 2)
                self.table.setColumnWidth(col, 300)
            elif col == 3 or col == 4:  # Customer column (index 3)
                self.table.setColumnWidth(col, 100)
            else:
                self.table.setColumnWidth(col, default_width)

        # Last column stretches to fill remaining space
        self.table.horizontalHeader().setStretchLastSection(True)

        # Visual polish
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Add to layout with stretch
        layout.addWidget(self.table, stretch=1)

        # Info label
        self.info_label = QLabel()
        layout.addWidget(self.info_label)

        # Buttons
        button_layout = QHBoxLayout()

        button_layout.addStretch()

        self.download_btn = QPushButton("Download Excel", objectName="SuccessButton")
        self.download_btn.clicked.connect(self.download_excel)
        button_layout.addWidget(self.download_btn)

        cancel_btn = QPushButton("Cancel", objectName="DangerButton")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def populate_months(self):
        """Populate month dropdown with available months in date range."""
        self.month_combo.addItem("All Months", None)

        current_date = QDate(self.date_from.year, self.date_from.month, 1)
        end_date = QDate(self.date_to.year, self.date_to.month, 1)

        months = []
        while current_date <= end_date:
            month_str = current_date.toString("MMMM yyyy")
            month_value = (current_date.year(), current_date.month())
            months.append((month_str, month_value))
            current_date = current_date.addMonths(1)

        for month_str, month_value in months:
            self.month_combo.addItem(month_str, month_value)

    def load_data(self):
        """Load data from database."""
        try:
            self.full_data = db_call.get_export_data(self.date_from, self.date_to)
            self.apply_filter()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {str(e)}")
            self.reject()

    def apply_filter(self):
        """Apply month filter to data and update preview + generate in-memory Excel."""
        try:
            selected_month = self.month_combo.currentData()

            if selected_month is None:
                self.filtered_data = self.full_data
            else:
                year, month = selected_month
                self.filtered_data = []

                for row in self.full_data:
                    row_date = row[1]
                    if isinstance(row_date, str):
                        row_date = datetime.strptime(row_date, "%d-%m-%Y").date()

                    if hasattr(row_date, 'year'):
                        if row_date.year == year and row_date.month == month:
                            self.filtered_data.append(row)
                    elif isinstance(row_date, datetime):
                        if row_date.year == year and row_date.month == month:
                            self.filtered_data.append(row)

            # Generate in-memory Excel file after filtering
            self.generate_temp_excel()

            self.update_table()
        except Exception as e:
            print(f"Filter error: {e}")
            self.filtered_data = []
            self.excel_bytes = None
            self.update_table()

    def update_table(self):
        """Update table with filtered data."""
        self.table.setRowCount(len(self.filtered_data))

        for row_idx, row_data in enumerate(self.filtered_data):
            for col_idx, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                self.table.setItem(row_idx, col_idx, item)

        self.info_label.setText(f"Showing {len(self.filtered_data)} records")

    def generate_temp_excel(self):
        """Generate formatted Excel file in memory from current filtered_data."""
        if not self.filtered_data:
            self.excel_bytes = None
            return

        try:
            default_headers = ["F1_t_uid", "t_date", "t_customer", "T_prodcode", "t_matcode", "t_con", "F1_t_deleted"]
            df = pd.DataFrame(self.filtered_data, columns=default_headers)

            # Clean "Con" column
            con_col_idx = self.headers.index("Con")
            df.iloc[:, con_col_idx] = pd.to_numeric(
                df.iloc[:, con_col_idx].astype(str).str.replace(",", ""), errors='coerce'
            )

            # Create in-memory file
            buffer = io.BytesIO()

            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Formulas')
                worksheet = writer.sheets['Formulas']

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center')

                # Apply borders and alignment
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                               min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                        # Column 6 = "Con" → center align
                        if cell.column == 6:
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align

                # Header styling
                for cell in worksheet[1]:
                    cell.font = Font(bold=False)
                    cell.alignment = left_align

                # Auto-size columns
                for idx, col_name in enumerate(df.columns, 1):
                    max_len = max(
                        len(str(val)) if pd.notna(val) else 0
                        for val in df[col_name]
                    )
                    max_len = max(max_len, len(col_name))
                    adjusted_width = min(max_len + 2, 50)
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = adjusted_width

            buffer.seek(0)
            self.excel_bytes = buffer  # Store for later use (download/email)

        except Exception as e:
            print(f"Failed to generate temp Excel: {e}")
            self.excel_bytes = None

    def download_excel(self):
        """Download the pre-generated in-memory Excel file."""
        if not self.excel_bytes:
            QMessageBox.warning(self, "No Data", "No data to export. Try filtering first.")
            return

        default_filename = "prod_formula.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", default_filename, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            with open(file_path, "wb") as f:
                f.write(self.excel_bytes.getvalue())

            QMessageBox.information(
                self,
                "Export Successful",
                f"Exported <b>{len(self.filtered_data)}</b> records to:<br>{file_path}"
            )

            self.parent_widget.log_audit_trail(
                "Data Export",
                f"Exported formulation table to {file_path}"
            )
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to save file:\n{str(e)}")

    def send_to_email(self):
        if not self.excel_bytes or not self.excel_bytes.getvalue():
            QMessageBox.warning(self, "No Data", "No data to send. Please select a month first.")
            return

        # Optional: Let user input recipient
        from PyQt6.QtWidgets import QInputDialog
        recipient, ok = QInputDialog.getText(
            self, "Send Email", "Enter recipient email:", text="ppsycho109@gmail.com"
        )
        if not ok or not recipient.strip():
            return

        try:
            send_email_with_excel_bytes(
                excel_bytes=self.excel_bytes.getvalue(),
                filename="prod_formula.xlsx",
                recipient_email=recipient.strip(),
                subject=f"Monthly Report",
                body="Please find the attached Excel report."
            )

            QMessageBox.information(self, "Success", f"Email sent successfully to:\n{recipient}")
            self.parent_widget.log_audit_trail("Email Export", f"Sent formula export to {recipient}")

        except Exception as e:
            QMessageBox.critical(self, "Email Failed", f"Could not send email:\n{str(e)}")
